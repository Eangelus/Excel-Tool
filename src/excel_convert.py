# Bernecker Thomas
# klasse zum laden einer excel und zurück geben in deiner liste
# logic funktion zum durchsuchen an text in der excel
from datetime import datetime

from openpyxl import Workbook, load_workbook
from src.appSetting import appSettings
import os


class excelImport():
    # golbale Variablen
    pfad = ""
    dictonary = {}
    indexZumAnzeigen =  []
    appSeting = appSettings()
    #############################################
    def __init__(self):
        # Kontruktor

        ## Lade die settings
        self.appSeting = appSettings()
        self.appSeting.loadSettings()

        print(self.appSeting.settingList["ExcelSettings"]["ImportVerzeichnis"] + "\\" +self.appSeting.settingList["ExcelSettings"]["ImportName"])
        ## Erstelle verzeichniss fürs die backupdatein wen es nicht vorhanden ist
        print("Erstelle Verzeichniss")
        if not os.path.exists(self.appSeting.settingList["ExcelSettings"]["UnterOrdnerErstellung"]):
            os.makedirs(self.appSeting.settingList["ExcelSettings"]["UnterOrdnerErstellung"])

        print("erstelle Kopie in den neuen Odner")
        ## Importiere die zu laden datei aus der cloud / netzwerk
        self.excelImport()
        ## excel in den speicher geladen..
        print("kopieren abgeschlossen")

        self.wb = load_workbook(self.appSeting.settingList["ExcelSettings"]["ImportVerzeichnis"] + "\\" + self.appSeting.settingList["ExcelSettings"]["ImportName"] + self.appSeting.settingList["ExcelSettings"]["DateiEndung"])
        ## Ecxel speicher auf localen datenträger

        self.ws = self.wb.active
        #self.wb.close() #### ---------------- muss?
        self.ws2 = self.wb[self.appSeting.settingList["ExcelSettings"]["Worksheet_2"]]

    # funktion zum laden der excel

    def loadExcel(self):
        self.appSeting.loadSettings()

        self.wb = load_workbook(self.appSeting.settingList["ExcelSettings"]["ImportVerzeichnis"] + "\\" +self.appSeting.settingList["ExcelSettings"]["ImportName"]  + self.appSeting.settingList["ExcelSettings"]["DateiEndung"])
        self.ws = self.wb.active
        self.ws2 = self.wb[self.appSeting.settingList["ExcelSettings"]["Worksheet_2"]]
        print("Laden abgeschlossen")

    #funktion zum suchen nach text in der ecxel
    def sucheNachName(self, a, spalte):
        ## locale varibale
        indexliste = []

        ## durchsuche jede spalte nach den text
        for element in self.ws[f'{spalte}']:

            name_in_der_Spalte = str(element.value)

            if a.lower() in name_in_der_Spalte.lower():
                index = element.row
                indexliste.append(index)
                if len(indexliste) == 0:
                    leerstellen_Platz = a.split()
                    zweiteKombo_desNames = leerstellen_Platz[1] + " " + leerstellen_Platz[0]
                    if zweiteKombo_desNames.lower() in name_in_der_Spalte.lower():
                        index = element.row
                        indexliste.append(index)
                        if zweiteKombo_desNames.lower() == name_in_der_Spalte.lower():
                            index = element.row
                            indexliste.append(index)
                        if a.lower() == name_in_der_Spalte.lower():
                            index = element.row
                            indexliste.append(index)
        # wenn eintrag gefunden ist dan zeilenummer in die liste
        indexliste = set(indexliste)
        indexliste = list(indexliste)

        return indexliste

    ## suche nach uid in bestimmter spalte
    def sucheNachUID(self, a, spalte):
        indexliste = []
        for element in self.ws[f'{spalte}']:
            name_in_der_Spalte = str(element.value)
                # wenn der übernommene text ( in klein ) den in der spalte enspricht
            if a.lower() in name_in_der_Spalte.lower():

                index = element.row

                indexliste.append(index)
            if a.lower() == name_in_der_Spalte.lower():
                index = element.row
                indexliste = []
                indexliste.append(index)
                    # wenn der übernommene text der in der spalte übereinstimmt nur ( vorname nachname vertauscht )

            indexliste = set(indexliste)
            indexliste = list(indexliste)
        #if indexliste != 0 and indexliste != None:
            ##   <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< TESTE auf kein bestelltes modell 19.12
        for element in indexliste:
            if self.ws[f'{self.appSeting.settingList["ExcelSettings"]["BestelltesModellSpalte"]}{element}'].value == None:
                indexliste.remove(element)
        return indexliste

        # funktion zum eintragen des context in der excel
    def saveContent(self,index, spalte, text):
        self.ws[f'{index}, {spalte}'] = str(text)
        ## speicher der ecxel datei
    def saveExcel(self, pfad, name, endung):
        self.wb.save(pfad+ "\\" +name+ f"_{datetime.today().strftime('%H %Y.%m.%d')}" + endung )#" "+ datetime.today().strftime('%H:%M:%S %Y.%m.%d') +
        print("Speichern erfolgt...")

        ## funktion zum beenden der app sicherheisthalber zum speiocher
    def exit(self):
        print("Beginne aufräumen..")
        print("Speicher nochmal in Excel ab..")
        try:

            self.wb.save(self.appSeting.settingList["ExcelSettings"]["BackupOrt"] + "\\" +
                         self.appSeting.settingList["ExcelSettings"]["BackupName"] + f" {datetime.today().strftime('%H %Y.%m.%d')}" +
                         self.appSeting.settingList["ExcelSettings"]["DateiEndung"])
            print("Speicherung erfolgreich!")
        except:
            print("Speicher FEHLGESCHLAGEN ! Datei war offen!")
        self.wb.close()
        print("Fertig geputzt")

        ## funktion zum laden der externen datei und speichern der datei im verzeichniss
    def excelImport(self):

        print("Erstelle Backup...")
        self.wb = load_workbook(self.appSeting.settingList["ExcelSettings"]["ImportVerzeichnis"] + "\\"+ self.appSeting.settingList["ExcelSettings"]["ImportName"]+ self.appSeting.settingList["ExcelSettings"]["DateiEndung"]  )
        print("Import erfolgreich..")
        self.ws = self.wb.active

        self.ws2 = self.wb[self.appSeting.settingList["ExcelSettings"]["Worksheet_2"]]

        self.wb.save(self.appSeting.settingList["ExcelSettings"]["BackupOrt"] + "\\" +self.appSeting.settingList["ExcelSettings"]["BackupName"] + self.appSeting.settingList["ExcelSettings"]["DateiEndung"])
        self.wb.close()
        print("speichern abgeschlossen")

    def excelExport(self, ExportOrt, ExportName, endung):
        self.wb.save(ExportOrt+"\\" +ExportName+endung)
