from datetime import datetime

from openpyxl import Workbook, load_workbook

import src.appSetting

## Bernecker Thomas Class zum exportieren der Sn und checken
# hierbei wird eine 2 excel geöffnet und werte übertragen   Übersicht Hardware
# 1.12.2022
#


import os

class snExport():

    def __init__(self):
        self.appSettings = src.appSetting.appSettings()

        ## importiere von extern die excel
        self.appSettings.loadSettings()
        if not os.path.exists(self.appSettings.settingList["ExcelSettings"]["UnterOrdnerErstellung"]):
            os.makedirs(self.appSettings.settingList["ExcelSettings"]["UnterOrdnerErstellung"])
        print("Lade... " + f'{self.appSettings.settingList["ExcelSettings"]["LadeLieferueberSicht"]}' )
        self.wb = load_workbook(self.appSettings.settingList["ExcelSettings"]["LadeLieferueberSicht"] )
        print("Speichere ab unter:" + f'{self.appSettings.settingList["ExcelSettings"]["BackupOrt"]}' + "\\"+f'{self.appSettings.settingList["ExcelSettings"]["BackupNameLiefer"]}'+f'{self.appSettings.settingList["ExcelSettings"]["DateiEndung"]}' )
        self.wb.save(f'{self.appSettings.settingList["ExcelSettings"]["BackupOrt"]}' + "\\"+f'{self.appSettings.settingList["ExcelSettings"]["BackupNameLiefer"]}'+f'{self.appSettings.settingList["ExcelSettings"]["DateiEndung"]}')
        self.wb.close()
    ## f. zum überprüfen und übertragen
    def check(self, inputModel ,inputUID,inputImg,inputHostname, inputSn, inputName,inputKostenstelle, inputHinweis, inputDatum):

        hostname = inputHostname
        modellspalte = self.appSettings.settingList["Excel2Einstellungen"]["Eintrag1"]
        modell = inputModel
        InputSn = inputSn
        toCheckSpalte = self.appSettings.settingList["Excel2Einstellungen"]["Prüfspalte"]
        nachnameVari = inputName
        uidVari = inputUID
        hinweissVari = inputHinweis
        kostenstelle = inputKostenstelle
        datum = inputDatum
        img = inputImg
        print("start search")
        print("Lade: " + self.appSettings.settingList["ExcelSettings"]['LadeLieferueberSicht'])
        wb = load_workbook(self.appSettings.settingList["ExcelSettings"]['LadeLieferueberSicht'])
        ws = wb.active

        for element in ws[f'{modellspalte}']:
            print("looking up for ")
            print(element.value)
            if element.value == modell:
                if ws[f'{toCheckSpalte}{element.row}'].value == None:
                    print("match and set" + str(element.row))
                    ws[f'{toCheckSpalte}{element.row}'] = InputSn
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag2"]}{element.row}'] = hostname
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag3"]}{element.row}'] = kostenstelle
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag4"]}{element.row}'] = nachnameVari
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag5"]}{element.row}'] = uidVari
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag6"]}{element.row}'] = hinweissVari
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag7"]}{element.row}'] = datum
                    ws[f'{self.appSettings.settingList["Excel2Einstellungen"]["Eintrag8"]}{element.row}'] = img
                    break

        wb.save(f'{self.appSettings.settingList["ExcelSettings"]["BackupOrt"]}' + "\\"+f'{self.appSettings.settingList["ExcelSettings"]["BackupNameLiefer"]}'+ f"_{datetime.today().strftime('%H %Y.%m.%d')}"+f'{self.appSettings.settingList["ExcelSettings"]["DateiEndung"]}')
        wb.save(self.appSettings.settingList["ExcelSettings"]['LadeLieferueberSicht'])
        wb.close()

    def ExportLiefer(self):
        wb = load_workbook(self.appSettings.settingList["ExcelSettings"]['LadeLieferueberSicht'])
        wb.save(self.appSettings.settingList["ExcelSettings"]['LadeLieferueberSicht'])
        wb.save(self.appSettings.settingList["ExcelSettings"]['BackupOrt']+"\\"+self.appSettings.settingList["ExcelSettings"]['BackupNameLiefer'])